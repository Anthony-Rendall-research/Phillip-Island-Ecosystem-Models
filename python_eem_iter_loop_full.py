from openpyxl import load_workbook
from string import ascii_lowercase
import itertools
import numpy as np
import numpy.matlib
from scipy.integrate import ode
import copy
import pickle
import matplotlib.pyplot as plt
import copy


import pandas as pd



def read_matrix(fname):
    wb = load_workbook(filename=fname)
    sheets = wb.sheetnames
    ws = wb[sheets[0]]
    n_row = num_rows(ws)
    n_col = num_cols(ws)
    if n_row != n_col:
        raise Exception('Number of rows and columns is not consistant')
    m = np.array(read_all_rows(ws, 2, n_row, 2, n_col))
    return m


def read_vector(fname):
    wb = load_workbook(filename=fname)
    sheets = wb.sheetnames
    ws = wb[sheets[0]]
    num_row = num_rows(ws, st_pt=1)
    output = np.array(read_all_rows(ws, 1, num_row, 2, 2))
    return output


def read_row(ws, row, st_pt, ed_pt):
    counter = 1
    output = []
    for s in iter_all_strings():
        if counter >= st_pt:
            output.append(ws[s + str(row)].value)
        if counter == ed_pt:
            return output
        counter += 1


def read_all_rows(ws, rows, rowe, st_pt, ed_pt):
    arr = []
    for row in range(rows, rowe + 1):
        c_row = read_row(ws, row, st_pt, ed_pt)
        c_row = none_2_zero(c_row)
        arr.append(c_row)
    return arr


def none_2_zero(arr):
    for i in range(len(arr)):
        if arr[i] is None:
            arr[i] = 0
    return arr


def num_rows(ws, st_pt=2):
    flag = True
    while flag:
        if ws['A' + str(st_pt)].value is None:
            return st_pt - 1
        else:
            st_pt += 1


def num_cols(ws, st_pt=2):
    count = 1
    for s in iter_all_strings():
        if count < st_pt:
            count += 1
            continue
        if ws[s + str(1)].value is None:
            return count - 1
        else:
            count += 1


def iter_all_strings():
    size = 1
    while True:
        for s in itertools.product(ascii_lowercase, repeat=size):
            yield "".join((s.upper() for s in s))
        size += 1


def equilibrium_state(amatrix, r_input):
    ia = np.linalg.inv(amatrix)
    n_eq = -np.matmul(ia, r_input)
    return n_eq


def calc_jacobian(A, r, n):
    i_matrix = np.eye(len(n))
    if len(n.shape) == 1:
        n_array = np.matlib.repmat(n, len(n), 1).T
    else:
        n_array = np.matlib.repmat(n, 1, len(n))
    J = i_matrix * r + A * n_array + i_matrix * np.matmul(A, n)

    return J


def calc_stability(A, r, n):
    J = calc_jacobian(A, r, n)
    ev = np.real(np.linalg.eig(J)[0])
    max_eig = np.max(ev)
    if max_eig < 0:
        return True
    else:
        return False


def draw_parameters(A, r):
    int_str = np.random.uniform(0, 1, np.shape(A))
    unc_int = np.abs(A) == 2
    unc_draws = np.random.uniform(0, 1, np.shape(A))
    A[(unc_draws < .5) & unc_int] = 0
    A[A == 2] = 1
    A[A == -2] = -1
    A_vals = (A * int_str)
    # A_vals = multiply_diag(A_vals)
    A_vals -= np.eye(np.shape(A_vals)[0])
    r_vals = np.random.uniform(0, 1, np.shape(r))
    return A_vals, r_vals

overall_count = []

def gen_stable_param_set(A, r):
    flag = 0
    count = 0
    while flag == 0:
        At, rt = draw_parameters(A, r)
        n = equilibrium_state(At, rt)
        if np.all(n > 0):
            st = calc_stability(At, rt, n)
            if st:
                count += 1
                overall_count.append(count)
                return At, rt, n
        count += 1


def multiply_diag(a, factor):
    a[np.eye(a.shape[0]) == 1] = a[np.eye(a.shape[0]) == 1] * factor
    return a


def remove_row_col(a_matrix, ind):
    a_shp = a_matrix.shape[0]
    if ind == 0:
        return a_matrix[1:a_shp, 1:a_shp]
    if ind == a_shp - 1:
        return a_matrix[0:a_shp - 1, 1:a_shp - 1]
    # If it is a 'central' option
    a_top_left = a_matrix[0:ind, 0:ind]
    a_top_right = a_matrix[0:ind, ind + 1:a_shp]
    a_bottom_left = a_matrix[ind + 1:a_shp, 0:ind]
    a_bottom_right = a_matrix[ind + 1:a_shp, ind + 1:a_shp]
    a_top = np.hstack([a_top_left, a_top_right])
    a_bottom = np.hstack([a_bottom_left, a_bottom_right])
    a_new = np.vstack([a_top, a_bottom])
    return a_new


def remove_rows_cols(A, r, inds):
    inds.sort()
    inds.reverse()
    An = A
    for ind in inds:
        An = remove_row_col(An, ind)
    rn = np.delete(r, inds)
    return An, rn


def add_rows_cols(A, r, n, inds, value=0):
    inds.sort()
    for ind in inds:
        A = np.insert(A, ind, value, axis=0)
        A = np.insert(A, ind, value, axis=1)
        r = np.insert(r, ind, value)
        n = np.insert(n, ind, value)
    return A, r, n

def gen_other_params(Ap, A_input, rp):
    A_draw, r_draw = draw_parameters(A_input, rp)
    locs = (A_input != 0) == (Ap == 0)
    rlocs = rp == 0
    Ap[locs] = A_draw[locs]
    rp[rlocs] = r_draw[rlocs]
    return Ap, rp

def gen_reduced_params(A_input, r, inds=None, reps=1):
    # inds is the species nums to remove
    A_output, r_output, n_output = [], [], []
    if inds is not None:
        A, r = remove_rows_cols(A_input, r, inds)
    else:
        A = A_input
    for i in range(reps):
        Ap, rp, Np = gen_stable_param_set(A, r)
        if inds is not None:
            Ap, rp, Np = add_rows_cols(Ap, rp, Np, inds)
            Ap, rp = gen_other_params(Ap, A_input, rp)
        A_output.append(Ap)
        r_output.append(rp)
        n_output.append(Np)
    A_output = np.array(A_output)
    r_output = np.array(r_output)
    n_output = np.array(n_output)
    return A_output, r_output, n_output


def de(t, y, A, r):
    return r * y + np.matmul(A, y) * y


def de_solve(T, y, A, r):
    rde = ode(de).set_integrator('lsoda', method='bdf', with_jacobian=False)
    rde.set_initial_value(y, 0).set_f_params(A, r)
    return np.real(rde.integrate(T))

def de_fix(t, y, A, r, f_id):
    rates = r * y + np.matmul(A, y) * y
    rates[f_id] = 0
    return rates


def de_solve_fix(T, y, A, r, f_id):
    rde = ode(de_fix).set_integrator('lsoda', method='bdf', with_jacobian=False)
    rde.set_initial_value(y, 0).set_f_params(A, r, f_id)
    return np.real(rde.integrate(T))


class EEM:
    def __init__(self, a, r, rem=None, max_sets=np.inf):
        self.a = a
        self.r = r
        self.rem = rem
        self.max = max_sets
        self.counter = 0

    def __iter__(self):
        return self

    def __next__(self):
        if self.counter >= self.max:
            raise StopIteration
        Ap, rp, Np = gen_reduced_params(self.a, self.r, self.rem, 1)
        self.counter += 1
        return Ap[0], rp[0], Np[0]

class EEM_rem:
    def __init__(self, EEM_gen, removal, response, max_iter=np.inf):
        self.EEM_gen = EEM_gen
        self.rem = removal
        self.resp = np.array(response)
        self.count = 0
        self.max = max_iter

    def __iter__(self):
        iter(self.EEM_gen)
        return self

    def __next__(self):
        if self.count >= self.max:
            raise StopIteration
        flag = 0
        c = 0
        while flag == 0:
            c += 1
            # print(c)
            a, r, n = next(self.EEM_gen)
            n_change = copy.copy(n)
            n_change[self.rem] = 0
            n_new = de_solve(.1, n_change, a, r)
            change = n_new > n
            # print(change[self.resp[:,0]])
            if (change[self.resp[:, 0]] == self.resp[:, 1]).all():
                flag = 1
        self.count += 1
        # print(a, r, n)
        return a, r, n

class EEM_stable_from_prev_conds:
    def __init__(self, EEM_gen, removal, max_iter=np.inf, return_num_unstable=False):
        self.EEM_gen = EEM_gen
        self.rem = removal
        self.count = 0
        self.max = max_iter
        self.return_unstable = return_num_unstable

    def __iter__(self):
        iter(self.EEM_gen)
        return self

    def __next__(self):
        if self.count >= self.max:
            raise StopIteration
        flag = 0
        num_unstable = 0
        while flag == 0:
            a, r, n = next(self.EEM_gen)
            a_rem, r_rem = remove_rows_cols(a, r, self.rem)
            n_rem = equilibrium_state(a_rem, r_rem)
            if calc_stability(a_rem, r_rem, n_rem):
                for ind in np.sort(self.rem): #reinsert the appropriate zeros
                    n_rem = np.insert(n_rem, ind, 0)
                self.count += 1
                if self.return_unstable:
                    return a, r, n_rem, num_unstable
                else:
                    return a, r, n_rem
            elif self.return_unstable:
                num_unstable += 1

class EEM_reintro:
    def __init__(self, ensemble, reintro_id, control_id, control_amount, T = 2):#number of years over which predictions occur
        self.ensemble = ensemble
        self.reintro = reintro_id
        self.control = control_id
        self.control_level = control_amount
        self.max = len(ensemble)
        self.count = 0
        self.outcomes = []
        self.T = T

    def __iter__(self):
        return self


    def __next__(self):
        if self.count >= self.max:
            raise StopIteration
        ratio = self.get_outcome(self.count)
        self.outcomes.append(ratio)
        self.count += 1
        return ratio

    def get_outcome(self, pset):
        if pset < len(self.outcomes):
            return self.outcomes[pset]
        param_set = self.ensemble[pset]
        a = param_set[0]
        r = param_set[1]
        n = param_set[2]
        n_old = copy.copy(n)
        start_abund = np.min(n[n != 0])
        if self.reintro is not None:
            n[self.reintro] = start_abund
        n_old = copy.copy(n)
        if len(self.control) > 0:
            n[self.control] = n[self.control]*self.control_level
        n_new = de_solve_fix(self.T, n, a, r, self.control)
        ratio = n_new/n_old
        ratio[np.isnan(ratio)] = 0
        return ratio




def save_object(obj, filename):
    with open(filename, 'wb') as output:  # Overwrites any existing file.
        pickle.dump(obj, output, pickle.HIGHEST_PROTOCOL)

def load_object(filename):
    with open(filename, 'rb') as input:
        return pickle.load(input)

def generate_phillip_island_ensemble(fname='PI_EEM', rep=10000, response = [], rem_node = [], final_removed=[], return_num_unstable=False):
    a_input = read_matrix('Phillip_islands_community_nominorints.xlsx').transpose()
    r_input = read_vector('Phillip_islands_r.xlsx')
    gen_init_stable = EEM(a_input, r_input, [2, 5])
    gen_cond_params = EEM_rem(gen_init_stable, rem_node, response)

    gen_final_param_set = EEM_stable_from_prev_conds(gen_cond_params, final_removed, max_iter=rep, return_num_unstable=return_num_unstable)
    if return_num_unstable:
        num_unstable = [params[3] for params in gen_final_param_set]
        return num_unstable
    else:
        param_sets = [[params[0], params[1], params[2]] for params in gen_final_param_set]
        fname = fname + '_' + str(rep) + '.pkl'
        save_object(param_sets, fname)

#reps = 10000
#rem_node = [0,16] # node to be removed
#response = [[7, True], [10, True]] # these increase when node is removed
#final_removed = [0] # return an ensemble without these species
#generate_phillip_island_ensemble(rep=reps, response = [[7, True], [10, True]], rem_node = [0],final_removed = [0, 16])
#ensemble = load_object('PI_EEM_{0}.pkl'.format(reps))
#control = [1]  # controlling cats
#control_level = 1.5 # control cats to 50% of current
#reintro_sp = None  # reintroduce bandicoots
#reintro = EEM_reintro(ensemble, reintro_sp, control, control_level)


if __name__ == "__main__":
    run_standard_analysis = False
    est_num_generated_param_sets = False
    generate_param_histograms = True
    uncerainty_across_param_sets_flag = False


##### START - CODE TO ESTIMATE NUMBER OF PARAMETER SETS DRAWN #####

    if est_num_generated_param_sets:
        results = np.empty(22)
        for i in range(0, 22):
            reps = 10000
            rem_node = [0, 16]  # node to be removed
            response = [[7, True], [10, True]]  # these increase when node is removed
            final_removed = [0]  # return an ensemble without these species

            #  num unstable is the number of unstable matrices at the 2nd filtering stage.
            num_unstable = generate_phillip_island_ensemble(rep=reps, response = [[7, True], [10, True]], rem_node = [0],final_removed = [0, 16], return_num_unstable=True)
            print("Total number of parameter sets generated: {total_number}, to get {req} acceptable\n\n"
                  "Percentage of acceptable param sets: {perc}".format(
                total_number=sum(overall_count),
                req=reps,
            perc=reps/sum(overall_count))
            )

#         Sample output:
#             "Total number of parameter sets generated: 192661, to get 10000 acceptable
# Percentage of acceptable param sets: 0.051904640793933386"

##### END - CODE TO ESTIMATE NUMBER OF PARAMETER SETS DRAWN #####


##### CODE TO RUN NORMAL ANALYSIS #####

    if run_standard_analysis:
            results = np.empty(22)
            for i in range(0, 22):
                reps = 10000
                rem_node = [0,16] # node to be removed
                response = [[7, True], [10, True]] # these increase when node is removed
                final_removed = [0] # return an ensemble without these species
                #generate_phillip_island_ensemble(rep=reps, response = [[7, True], [10, True]], rem_node = [0],final_removed = [0, 16])
                ensemble = load_object('PI_EEM_{0}.pkl'.format(reps))
                control = [1]  # controlling cats
                control_level = 0.25 # control cats to 50% of current
                reintro_sp = [16]  # reintroduce bandicoots
                reintro = EEM_reintro(ensemble, reintro_sp, control, control_level)
                temp_data = [abund[i] for abund in reintro]
                temp_result = np.mean(np.array(temp_data)>1)
                results[i] = temp_result

            #plt.hist(bandicoot_data, bins= 'auto')
            #plt.title("Frequency of increase: {}".format((int(mean_increase * 1000)) / 1000))
            #plt.show()

            # Print results out
            np.savetxt('Final_results_Bandicoot_introduction_2yrs_cat_supp_0_25.csv', results, delimiter = ',')


    # import matplotlib.pyplot as plt
    # from matplotlib.ticker import FuncFormatter

    # plt.style.use('seaborn-dark')
    # plt.rcParams.update({'figure.autolayout': True})
    # fig, ax = plt.subplots(figsize=(15, 8))
    # ax.bar(group_names, group_data)
    # labels = ax.get_xticklabels()
    # plt.setp(labels, rotation=45, horizontalalignment='right')
    # ax.set(ylim=[0, 1.1], ylabel='Frequency of increase')
    # plt.show()

    ##### CODE TO Look at uncertainy acros param sets #####

    if uncerainty_across_param_sets_flag:
        num_repeats = 100
        uncertainty_list = []
        for j in range(num_repeats):
            results = np.empty(22)
            print(f'running {j}')
            reps = 10000
            rem_node = [0, 16]  # node to be removed
            response = [[7, True], [10, True]]  # these increase when node is removed
            final_removed = [0]  # return an ensemble without these species
            generate_phillip_island_ensemble(rep=reps, response = [[7, True], [10, True]], rem_node = [0],final_removed = [0, 16])
            ensemble = load_object('PI_EEM_{0}.pkl'.format(reps))
            control = [1]  # controlling cats
            control_level = 0.25  # control cats to 50% of current
            reintro_sp = [16]  # reintroduce bandicoots
            reintro = EEM_reintro(ensemble, reintro_sp, control, control_level)
            reintro_results_list = list(reintro)
            for i in range(0, 22):
                temp_data = [abund[i] for abund in reintro_results_list]
                temp_result = np.mean(np.array(temp_data) > 1)
                results[i] = temp_result

            uncertainty_list.append(results)
        uncertainty_array = np.array(uncertainty_list)
        uncertainty_DF = pd.DataFrame(uncertainty_array)
        uncertainty_DF.to_csv('uncertainty_across_param_sets.csv', index=False)




    ##### NEW CODE TO GENERATE HISTOGRAMS #####

    if generate_param_histograms:
        reps = 10000
        ensemble = load_object('PI_EEM_{0}.pkl'.format(reps))
        a_input = read_matrix('Phillip_islands_community_nominorints.xlsx').transpose()
        r_input = read_vector('Phillip_islands_r.xlsx')

        non_zero_a = a_input != 0

        index_values = np.where(non_zero_a)
        a_param_names = [f'a_{int(i+1)}_{int(j)}' for i, j in zip(*index_values)]
        a_values_dict = {name: [] for name in a_param_names}

        r_param_names = ['r{}'.format(int(i+1)) for i in range(len(r_input))]
        r_values_dict = {name: [] for name in r_param_names}

        count = 0
        for a_matrix, r_matrix, n in ensemble:
            a_vals = a_matrix[non_zero_a]
            for a_val, val_name in zip(a_vals, a_param_names):
                if a_val != 0:
                    a_values_dict[val_name].append(a_val)
            for r_val, val_name in zip(r_matrix, r_param_names):
                r_values_dict[val_name].append(r_val)

        for a_name in a_param_names:
            current_values = a_values_dict[a_name]
            plt.hist(current_values)
            figure_name = '{name}_histogram'.format(name=a_name)
            plt.title(figure_name)
            plt.xlabel('Parameter value')
            plt.ylabel('Frequency')
            plt.savefig('param_dist_figures/{}.png'.format(figure_name))
            plt.close()

        for r_name in r_param_names:
            current_values = r_values_dict[r_name]
            plt.hist(current_values)
            figure_name = '{name}_histogram'.format(name=r_name)
            plt.title(figure_name)
            plt.xlabel('Parameter value')
            plt.ylabel('Frequency')
            plt.savefig('param_dist_figures/{}.png'.format(figure_name))
            plt.close()